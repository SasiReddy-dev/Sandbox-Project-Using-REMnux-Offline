import hashlib, json, os, time
from analyzer import analyze_file, simulate_dynamic_analysis
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)

UPLOAD_FOLDER = "uploads"
REPORT_FOLDER = "reports"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(REPORT_FOLDER, exist_ok=True)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/report")
def report_page():
    return render_template("report.html")

@app.route("/upload", methods=["POST"])
def upload():
    """
    Static analysis only:
    - save file
    - run analyze_file()
    - save JSON + Excel
    - return report + Excel link
    """
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file provided"}), 400

    filename = file.filename
    file_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(file_path)

    # Perform static analysis
    report = analyze_file(file_path)

    # Save JSON backup
    json_path = os.path.join(REPORT_FOLDER, filename + ".json")
    with open(json_path, "w") as f:
        json.dump(report, f, indent=2)

    # Create Excel report
    wb = Workbook()
    ws = wb.active
    ws.title = "File Analysis Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E2A78", end_color="1E2A78", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    ws.append(["Field", "Value"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for key, value in report.items():
        ws.append([key, str(value)])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[column].width = max_length + 2

    excel_path = os.path.join(REPORT_FOLDER, filename + ".xlsx")
    wb.save(excel_path)

    return jsonify({
        "report": report,
        "excel_path": f"/download/{filename}.xlsx"
    })


@app.route("/dynamic_simulate", methods=["POST"])
def dynamic_simulate():
    """
    Dynamic simulation only:
    - save file
    - run simulate_dynamic_analysis()
    - save dynamic JSON log
    - return dynamic log
    """
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file provided"}), 400

    filename = file.filename
    saved_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(saved_path)

    # Run safe dynamic simulation
    dynamic_log = simulate_dynamic_analysis(saved_path)

    # Save dynamic log
    dyn_path = os.path.join(REPORT_FOLDER, filename + ".dynamic.json")
    with open(dyn_path, "w") as fh:
        json.dump(dynamic_log, fh, indent=2)

    return jsonify({
        "dynamic_log": dynamic_log,
        "dynamic_saved": dyn_path
    })


@app.route("/download/<path:filename>")
def download(filename):
    file_path = os.path.join(REPORT_FOLDER, filename)
    if not os.path.exists(file_path):
        return jsonify({"error": "File not found"}), 404
    return send_file(file_path, as_attachment=True)


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=True)
